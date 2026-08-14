# 0_create_db.py
# DB 생성, 테이블 생성, CSV/JSON/XLSX 데이터 적재

import csv
from datetime import date, datetime
from io import BytesIO, StringIO
import json
from pathlib import Path
import sqlite3


DB_PATH = Path(__file__).with_name("fire_db.db")
REQUIRED_COLUMNS = {
    "marker_id",
    "location_x",
    "location_y",
    "manufacture_date",
    "pressure_status",
    "result",
}


def get_connection():
    return sqlite3.connect(DB_PATH)


def create_tables():
    # 시스템에서 사용하는 모든 테이블을 준비합니다.
    with get_connection() as connection:
        cursor = connection.cursor()
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS fire_extinguisher(
            marker_id INTEGER PRIMARY KEY,
            location_x REAL NOT NULL,
            location_y REAL NOT NULL,
            manufacture_date TEXT NOT NULL,
            last_inspection_date TEXT DEFAULT (datetime('now', '+9 hours')),
            pressure_status TEXT,
            result TEXT
        )
        """)
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS robot_battery_status(
            robot_name TEXT PRIMARY KEY,
            topic_name TEXT NOT NULL,
            percentage REAL,
            voltage REAL,
            current REAL,
            updated_at TEXT
        )
        """)
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS robot_dock_status(
            robot_name TEXT PRIMARY KEY,
            topic_name TEXT NOT NULL,
            dock_visible INTEGER,
            is_docked INTEGER,
            updated_at TEXT
        )
        """)
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS robot_camera_status(
            robot_name TEXT PRIMARY KEY,
            topic_name TEXT NOT NULL,
            image_filename TEXT,
            updated_at TEXT
        )
        """)
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS fire_extinguisher_inspection_history(
            inspection_id INTEGER PRIMARY KEY AUTOINCREMENT,
            marker_id INTEGER NOT NULL,
            robot_name TEXT NOT NULL,
            result TEXT NOT NULL,
            snapshot_filename TEXT,
            inspected_at TEXT NOT NULL DEFAULT (datetime('now', '+9 hours'))
        )
        """)
        history_columns = {
            row[1]
            for row in cursor.execute(
                "PRAGMA table_info(fire_extinguisher_inspection_history)"
            )
        }
        if "snapshot_filename" not in history_columns:
            cursor.execute(
                "ALTER TABLE fire_extinguisher_inspection_history "
                "ADD COLUMN snapshot_filename TEXT"
            )
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS system_metadata(
            metadata_key TEXT PRIMARY KEY,
            metadata_value TEXT NOT NULL,
            updated_at TEXT NOT NULL DEFAULT (datetime('now', '+9 hours'))
        )
        """)


def parse_optional_text(value):
    # 공백만 있는 문자열은 DB에 빈 문자열 대신 NULL로 저장합니다.
    if value is None:
        return None
    value = str(value).strip()
    return value if value else None


def normalize_date(value):
    # XLSX에서 날짜 셀이 date/datetime 객체로 읽히는 경우 ISO 문자열로 변환합니다.
    if isinstance(value, (date, datetime)):
        return value.date().isoformat() if isinstance(value, datetime) else value.isoformat()
    return str(value).strip()


def normalize_fire_extinguisher_entries(rows, fieldnames, source_name):
    # 파일 형식과 관계없이 공통 컬럼과 행 데이터를 검증합니다.
    if not fieldnames:
        raise ValueError(f"{source_name} 헤더가 없습니다.")

    missing_columns = REQUIRED_COLUMNS - set(fieldnames)
    if missing_columns:
        missing = ", ".join(sorted(missing_columns))
        raise ValueError(f"{source_name} 필수 컬럼이 없습니다: {missing}")

    entries = []
    for row_number, row in enumerate(rows, start=2):
        try:
            entries.append(
                {
                    "marker_id": int(row["marker_id"]),
                    "location_x": float(row["location_x"]),
                    "location_y": float(row["location_y"]),
                    "manufacture_date": normalize_date(row["manufacture_date"]),
                    "pressure_status": str(row["pressure_status"]).strip(),
                    "result": parse_optional_text(row["result"]),
                }
            )
        except (TypeError, ValueError, AttributeError) as error:
            raise ValueError(
                f"{source_name} {row_number}행 형식이 올바르지 않습니다: {error}"
            ) from error

    if not entries:
        raise ValueError(f"{source_name} 데이터가 없습니다.")

    return entries


def load_fire_extinguisher_entries(csv_text):
    # 업로드된 CSV를 서버 경로에 저장하지 않고 메모리에서 읽습니다.
    reader = csv.DictReader(StringIO(csv_text))
    return normalize_fire_extinguisher_entries(reader, reader.fieldnames, "CSV")


def load_fire_extinguisher_entries_from_json(file_bytes):
    try:
        payload = json.loads(file_bytes.decode("utf-8-sig"))
    except (UnicodeDecodeError, json.JSONDecodeError) as error:
        raise ValueError(f"JSON 파일 형식이 올바르지 않습니다: {error}") from error

    rows = payload.get("rows") if isinstance(payload, dict) else payload
    if not isinstance(rows, list) or not all(isinstance(row, dict) for row in rows):
        raise ValueError("JSON은 객체 목록 또는 rows 목록 형식이어야 합니다.")

    fieldnames = set().union(*(row.keys() for row in rows)) if rows else set()
    return normalize_fire_extinguisher_entries(rows, fieldnames, "JSON")


def load_fire_extinguisher_entries_from_xlsx(file_bytes):
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise ValueError("XLSX 지원을 위해 python3-openpyxl 패키지가 필요합니다.") from error

    try:
        workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        worksheet = workbook.active
        values = worksheet.iter_rows(values_only=True)
        header_row = next(values, None)
        fieldnames = [str(value).strip() if value is not None else "" for value in header_row or []]
        rows = [
            dict(zip(fieldnames, row))
            for row in values
            if any(value is not None for value in row)
        ]
    except Exception as error:
        raise ValueError(f"XLSX 파일을 읽을 수 없습니다: {error}") from error
    finally:
        if "workbook" in locals():
            workbook.close()

    return normalize_fire_extinguisher_entries(rows, fieldnames, "XLSX")


def overwrite_fire_extinguisher_entries(entries):
    # 검증이 끝난 항목만 하나의 트랜잭션으로 DB에 반영합니다.
    with get_connection() as connection:
        cursor = connection.cursor()
        cursor.execute("DELETE FROM fire_extinguisher")
        cursor.executemany(
            """
            INSERT INTO fire_extinguisher (
                marker_id,
                location_x,
                location_y,
                manufacture_date,
                pressure_status,
                result
            )
            VALUES (
                :marker_id,
                :location_x,
                :location_y,
                :manufacture_date,
                :pressure_status,
                :result
            )
            """,
            entries,
        )


def save_import_filename(filename):
    # 브라우저가 보낸 경로를 제외하고 원본 파일명만 저장합니다.
    safe_filename = Path(filename).name
    with get_connection() as connection:
        connection.execute(
            """
            INSERT INTO system_metadata (metadata_key, metadata_value, updated_at)
            VALUES ('last_import_filename', ?, datetime('now', '+9 hours'))
            ON CONFLICT(metadata_key) DO UPDATE SET
                metadata_value = excluded.metadata_value,
                updated_at = excluded.updated_at
            """,
            (safe_filename,),
        )


def import_fire_extinguishers_from_csv(csv_bytes):
    # 웹 첨부 파일의 UTF-8 BOM을 허용하고 CSV 데이터를 적재합니다.
    try:
        csv_text = csv_bytes.decode("utf-8-sig")
    except UnicodeDecodeError as error:
        raise ValueError("CSV 파일은 UTF-8 인코딩이어야 합니다.") from error

    entries = load_fire_extinguisher_entries(csv_text)
    create_tables()
    overwrite_fire_extinguisher_entries(entries)
    return len(entries)


def import_fire_extinguishers(file_bytes, filename):
    # 업로드 파일 확장자에 맞는 파서를 선택한 뒤 공통 DB 저장 로직을 사용합니다.
    extension = Path(filename).suffix.lower()
    if extension == ".csv":
        try:
            entries = load_fire_extinguisher_entries(file_bytes.decode("utf-8-sig"))
        except UnicodeDecodeError as error:
            raise ValueError("CSV 파일은 UTF-8 인코딩이어야 합니다.") from error
    elif extension == ".json":
        entries = load_fire_extinguisher_entries_from_json(file_bytes)
    elif extension == ".xlsx":
        entries = load_fire_extinguisher_entries_from_xlsx(file_bytes)
    else:
        raise ValueError("CSV, JSON, XLSX 파일만 업로드할 수 있습니다.")

    create_tables()
    overwrite_fire_extinguisher_entries(entries)
    save_import_filename(filename)
    return len(entries)


def main():
    create_tables()
    print(f"DB 생성 완료: {DB_PATH}")


if __name__ == "__main__":
    main()
